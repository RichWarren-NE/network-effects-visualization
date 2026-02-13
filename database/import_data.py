"""
Network Effects: Film Festival Circulation Database - Data Import Script

Reads Encounters_Data.xlsx and imports normalised data into Supabase PostgreSQL.
Includes fuzzy duplicate detection, multi-value field parsing, and ISO country codes.

Usage:
    python database/import_data.py
"""

import os
import sys
import uuid
import logging
from datetime import datetime
from collections import defaultdict

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv
from Levenshtein import ratio as levenshtein_ratio

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'Encounters_Data.xlsx')
SHEET_NAME = 'encounters_filme_alle_selektion'
REPORT_PATH = os.path.join(os.path.dirname(__file__), '..', 'reports', 'data_quality_report.txt')
SCHEMA_PATH = os.path.join(os.path.dirname(__file__), 'schema.sql')
DUPLICATE_THRESHOLD = 0.85  # Levenshtein similarity threshold for title matching

VALID_STATUSES = {
    'screening', 'won', 'nominated', 'special_mention',
    'honorable_mention', 'second_place', 'third_place'
}

# ISO 3166-1 alpha-2 country codes for the 80 countries found in the dataset
COUNTRY_CODES = {
    'Afghanistan': 'AF', 'Albania': 'AL', 'Argentina': 'AR', 'Australia': 'AU',
    'Austria': 'AT', 'Bangladesh': 'BD', 'Belgium': 'BE',
    'Bosnia and Herzegovina': 'BA', 'Brazil': 'BR', 'Bulgaria': 'BG',
    'Cambodia': 'KH', 'Canada': 'CA', 'Chile': 'CL', 'China': 'CN',
    'Colombia': 'CO', 'Croatia': 'HR', 'Cyprus': 'CY', 'Czechia': 'CZ',
    'Democratic Republic of the Congo': 'CD', 'Denmark': 'DK', 'Egypt': 'EG',
    'Estonia': 'EE', 'Faroe Islands': 'FO', 'Finland': 'FI', 'France': 'FR',
    'Germany': 'DE', 'Ghana': 'GH', 'Greece': 'GR', 'Hong Kong': 'HK',
    'Hungary': 'HU', 'Iceland': 'IS', 'India': 'IN', 'Indonesia': 'ID',
    'Iran': 'IR', 'Ireland': 'IE', 'Israel': 'IL', 'Italy': 'IT',
    'Japan': 'JP', 'Kazakhstan': 'KZ', 'Kenya': 'KE', 'Lebanon': 'LB',
    'Lithuania': 'LT', 'Malaysia': 'MY', 'Malta': 'MT', 'Mexico': 'MX',
    'Morocco': 'MA', 'Netherlands': 'NL', 'Nigeria': 'NG', 'Norway': 'NO',
    'Pakistan': 'PK', 'Palestine': 'PS', 'Panama': 'PA', 'Philippines': 'PH',
    'Poland': 'PL', 'Portugal': 'PT', 'Qatar': 'QA',
    'Republic of Kosovo': 'XK', 'Romania': 'RO', 'Russia': 'RU',
    'Serbia': 'RS', 'Singapore': 'SG', 'Slovakia': 'SK', 'Slovenia': 'SI',
    'Somalia': 'SO', 'South Africa': 'ZA', 'South Korea': 'KR', 'Spain': 'ES',
    'Sri Lanka': 'LK', 'Sudan': 'SD', 'Sweden': 'SE', 'Switzerland': 'CH',
    'Syrian Arab Republic': 'SY', 'Taiwan': 'TW', 'Thailand': 'TH',
    'Togo': 'TG', 'Turkey': 'TR', 'United Kingdom': 'GB',
    'United States': 'US', 'Vietnam': 'VN', 'Zimbabwe': 'ZW',
}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_int(val):
    """Convert a value to int, handling floats and strings."""
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert to stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def split_csv(val):
    """Split a comma-separated string into a list of trimmed, non-empty values."""
    if val is None:
        return []
    return [v.strip() for v in str(val).split(',') if v.strip()]


def get_connection():
    """Create a PostgreSQL connection from .env credentials."""
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

    host = os.getenv('SUPABASE_HOST')
    db = os.getenv('SUPABASE_DB')
    user = os.getenv('SUPABASE_USER')
    password = os.getenv('SUPABASE_PASSWORD')
    port = os.getenv('SUPABASE_PORT', '5432')

    if not all([host, db, user, password]):
        log.error("Missing Supabase credentials in .env file")
        sys.exit(1)

    log.info(f"Connecting to {host}:{port}/{db} as {user}...")
    try:
        conn = psycopg2.connect(
            host=host, database=db, user=user, password=password,
            port=port, sslmode='require',
            connect_timeout=30,
        )
        conn.autocommit = False
        log.info("Connected successfully")
        return conn
    except psycopg2.Error as e:
        log.error(f"Connection failed: {e}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# Schema creation
# ---------------------------------------------------------------------------

def create_schema(conn):
    """Execute the schema SQL file to create all tables."""
    log.info("Creating database schema...")
    with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
        schema_sql = f.read()

    cur = conn.cursor()
    try:
        cur.execute(schema_sql)
        conn.commit()
        log.info("Schema created successfully")
    except psycopg2.Error as e:
        conn.rollback()
        # If tables already exist, drop and recreate
        if 'already exists' in str(e):
            log.warning("Tables already exist. Dropping and recreating...")
            drop_all_tables(conn)
            cur = conn.cursor()
            cur.execute(schema_sql)
            conn.commit()
            log.info("Schema recreated successfully")
        else:
            log.error(f"Schema creation failed: {e}")
            raise
    finally:
        cur.close()


def drop_all_tables(conn):
    """Drop all project tables in correct order (respecting FK constraints)."""
    cur = conn.cursor()
    tables = [
        'film_genres', 'film_countries', 'film_directors',
        'screenings', 'genres', 'countries', 'directors', 'festivals', 'films',
    ]
    for table in tables:
        cur.execute(f"DROP TABLE IF EXISTS {table} CASCADE")
    # Drop the trigger function too
    cur.execute("DROP FUNCTION IF EXISTS update_updated_at_column() CASCADE")
    conn.commit()
    cur.close()
    log.info("Dropped existing tables")


# ---------------------------------------------------------------------------
# Data reading
# ---------------------------------------------------------------------------

def read_excel():
    """Read the Excel file and return a list of row dicts."""
    log.info(f"Reading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows.append(record)

    wb.close()
    log.info(f"Read {len(rows)} rows from Excel")
    return rows


# ---------------------------------------------------------------------------
# Fuzzy duplicate detection
# ---------------------------------------------------------------------------

def detect_duplicates(films_data):
    """
    Detect potential duplicate films using fuzzy matching.
    Matches by: similar title (>85%) + same year + same director.
    Returns list of potential duplicate pairs for manual review.
    """
    log.info("Running fuzzy duplicate detection...")
    duplicates = []
    film_keys = list(films_data.keys())

    for i in range(len(film_keys)):
        for j in range(i + 1, len(film_keys)):
            key_a = film_keys[i]
            key_b = film_keys[j]
            film_a = films_data[key_a]
            film_b = films_data[key_b]

            # Same year required
            if film_a['year'] != film_b['year']:
                continue
            if film_a['year'] is None:
                continue

            # Title similarity
            title_sim = levenshtein_ratio(
                film_a['title'].lower(),
                film_b['title'].lower()
            )
            if title_sim < DUPLICATE_THRESHOLD:
                continue

            # Check for shared director
            dirs_a = set(d.lower() for d in film_a['directors'])
            dirs_b = set(d.lower() for d in film_b['directors'])
            shared_directors = dirs_a & dirs_b

            if shared_directors or (not dirs_a and not dirs_b):
                duplicates.append({
                    'film_a_title': film_a['title'],
                    'film_b_title': film_b['title'],
                    'year': film_a['year'],
                    'title_similarity': round(title_sim * 100, 1),
                    'shared_directors': list(shared_directors) if shared_directors else ['(none)'],
                })

    log.info(f"Found {len(duplicates)} potential duplicate pairs")
    return duplicates


# ---------------------------------------------------------------------------
# Data extraction and normalisation
# ---------------------------------------------------------------------------

def extract_entities(rows):
    """
    Extract normalised entities from flat Excel rows.
    Returns dicts for films, festivals, directors, countries, genres, and screenings.
    """
    log.info("Extracting and normalising entities...")

    # Use (title, year, directors_str) as composite key for unique films
    films = {}       # key -> film dict
    festivals = {}   # festival_name -> festival dict
    directors = {}   # director_name -> director dict
    countries = {}   # country_name -> country dict
    genres = {}      # genre_name -> genre dict
    screenings = []  # list of screening dicts

    skipped_rows = 0

    for row_num, row in enumerate(rows, start=2):
        title = safe_str(row.get('title'))
        if not title:
            skipped_rows += 1
            continue

        original_title = safe_str(row.get('original_title'))
        year = safe_int(row.get('year'))
        duration = safe_int(row.get('duration'))
        synopsis = safe_str(row.get('synopsis'))
        industry_events_count = safe_int(row.get('industry_events_count'))
        genre_list = split_csv(row.get('genres'))
        country_list = split_csv(row.get('countries'))
        director_list = split_csv(row.get('directors'))
        event_type = safe_str(row.get('event_type'))
        event_name = safe_str(row.get('event_name'))
        event_year = safe_int(row.get('event_year'))
        category = safe_str(row.get('category'))
        status = safe_str(row.get('status'))

        # Validate status
        if status and status not in VALID_STATUSES:
            log.warning(f"Row {row_num}: Unknown status '{status}', setting to NULL")
            status = None

        # Build film key (title + year + sorted directors)
        directors_key = '|'.join(sorted(d.lower() for d in director_list))
        film_key = (title.lower(), year, directors_key)

        # Register film
        if film_key not in films:
            films[film_key] = {
                'film_id': str(uuid.uuid4()),
                'title': title,
                'original_title': original_title,
                'year': year,
                'duration': duration,
                'synopsis': synopsis,
                'industry_events_count': industry_events_count,
                'directors': director_list,
                'countries': country_list,
                'genres': genre_list,
            }
        else:
            # Update with non-null values if current record has more info
            existing = films[film_key]
            if not existing['original_title'] and original_title:
                existing['original_title'] = original_title
            if not existing['duration'] and duration:
                existing['duration'] = duration
            if not existing['synopsis'] and synopsis:
                existing['synopsis'] = synopsis

        # Register festival
        if event_name and event_name not in festivals:
            festivals[event_name] = {
                'festival_id': str(uuid.uuid4()),
                'festival_name': event_name,
                'event_type': event_type,
                'first_year': event_year,
                'last_year': event_year,
            }
        elif event_name and event_year:
            fest = festivals[event_name]
            if fest['first_year'] is None or (event_year and event_year < fest['first_year']):
                fest['first_year'] = event_year
            if fest['last_year'] is None or (event_year and event_year > fest['last_year']):
                fest['last_year'] = event_year

        # Register directors
        for d in director_list:
            if d not in directors:
                directors[d] = {
                    'director_id': str(uuid.uuid4()),
                    'director_name': d,
                }

        # Register countries
        for c in country_list:
            if c not in countries:
                countries[c] = {
                    'country_id': str(uuid.uuid4()),
                    'country_name': c,
                    'country_code': COUNTRY_CODES.get(c),
                }

        # Register genres
        for g in genre_list:
            if g not in genres:
                genres[g] = {
                    'genre_id': str(uuid.uuid4()),
                    'genre_name': g,
                }

        # Create screening record
        if event_name:
            screenings.append({
                'screening_id': str(uuid.uuid4()),
                'film_key': film_key,
                'festival_name': event_name,
                'event_year': event_year,
                'status': status,
                'category': category,
            })

    if skipped_rows:
        log.warning(f"Skipped {skipped_rows} rows with missing title")

    log.info(f"Extracted: {len(films)} films, {len(festivals)} festivals, "
             f"{len(directors)} directors, {len(countries)} countries, "
             f"{len(genres)} genres, {len(screenings)} screenings")

    return films, festivals, directors, countries, genres, screenings


# ---------------------------------------------------------------------------
# Database insertion
# ---------------------------------------------------------------------------

def insert_data(conn, films, festivals, directors, countries, genres, screenings):
    """Insert all normalised data into the database."""
    cur = conn.cursor()

    try:
        # 1. Insert films
        log.info(f"Inserting {len(films)} films...")
        film_values = [
            (f['film_id'], f['title'], f['original_title'], f['year'],
             f['duration'], f['synopsis'], f['industry_events_count'])
            for f in films.values()
        ]
        execute_values(cur,
            """INSERT INTO films (film_id, title, original_title, year, duration, synopsis, industry_events_count)
               VALUES %s ON CONFLICT DO NOTHING""",
            film_values,
            template="(%s, %s, %s, %s, %s, %s, %s)"
        )

        # 2. Insert festivals
        log.info(f"Inserting {len(festivals)} festivals...")
        festival_values = [
            (f['festival_id'], f['festival_name'], f['event_type'],
             f['first_year'], f['last_year'])
            for f in festivals.values()
        ]
        execute_values(cur,
            """INSERT INTO festivals (festival_id, festival_name, event_type, first_year, last_year)
               VALUES %s ON CONFLICT (festival_name) DO NOTHING""",
            festival_values,
            template="(%s, %s, %s, %s, %s)"
        )

        # 3. Insert directors
        log.info(f"Inserting {len(directors)} directors...")
        director_values = [
            (d['director_id'], d['director_name'])
            for d in directors.values()
        ]
        execute_values(cur,
            """INSERT INTO directors (director_id, director_name)
               VALUES %s ON CONFLICT (director_name) DO NOTHING""",
            director_values,
            template="(%s, %s)"
        )

        # 4. Insert countries
        log.info(f"Inserting {len(countries)} countries...")
        country_values = [
            (c['country_id'], c['country_name'], c['country_code'])
            for c in countries.values()
        ]
        execute_values(cur,
            """INSERT INTO countries (country_id, country_name, country_code)
               VALUES %s ON CONFLICT (country_name) DO NOTHING""",
            country_values,
            template="(%s, %s, %s)"
        )

        # 5. Insert genres
        log.info(f"Inserting {len(genres)} genres...")
        genre_values = [
            (g['genre_id'], g['genre_name'])
            for g in genres.values()
        ]
        execute_values(cur,
            """INSERT INTO genres (genre_id, genre_name)
               VALUES %s ON CONFLICT (genre_name) DO NOTHING""",
            genre_values,
            template="(%s, %s)"
        )

        # 6. Insert film_directors junction
        log.info("Inserting film-director relationships...")
        fd_values = []
        for film in films.values():
            for d_name in film['directors']:
                if d_name in directors:
                    fd_values.append((film['film_id'], directors[d_name]['director_id']))
        if fd_values:
            execute_values(cur,
                """INSERT INTO film_directors (film_id, director_id)
                   VALUES %s ON CONFLICT DO NOTHING""",
                fd_values,
                template="(%s, %s)"
            )
        log.info(f"  {len(fd_values)} film-director links")

        # 7. Insert film_countries junction
        log.info("Inserting film-country relationships...")
        fc_values = []
        for film in films.values():
            for c_name in film['countries']:
                if c_name in countries:
                    fc_values.append((film['film_id'], countries[c_name]['country_id']))
        if fc_values:
            execute_values(cur,
                """INSERT INTO film_countries (film_id, country_id)
                   VALUES %s ON CONFLICT DO NOTHING""",
                fc_values,
                template="(%s, %s)"
            )
        log.info(f"  {len(fc_values)} film-country links")

        # 8. Insert film_genres junction
        log.info("Inserting film-genre relationships...")
        fg_values = []
        for film in films.values():
            for g_name in film['genres']:
                if g_name in genres:
                    fg_values.append((film['film_id'], genres[g_name]['genre_id']))
        if fg_values:
            execute_values(cur,
                """INSERT INTO film_genres (film_id, genre_id)
                   VALUES %s ON CONFLICT DO NOTHING""",
                fg_values,
                template="(%s, %s)"
            )
        log.info(f"  {len(fg_values)} film-genre links")

        # 9. Insert screenings
        log.info(f"Inserting {len(screenings)} screenings...")
        screening_values = []
        for s in screenings:
            film_id = films[s['film_key']]['film_id']
            festival_id = festivals[s['festival_name']]['festival_id']
            screening_values.append((
                s['screening_id'], film_id, festival_id,
                s['event_year'], s['status'], s['category']
            ))
        if screening_values:
            execute_values(cur,
                """INSERT INTO screenings (screening_id, film_id, festival_id, event_year, status, category)
                   VALUES %s ON CONFLICT DO NOTHING""",
                screening_values,
                template="(%s, %s, %s, %s, %s, %s)"
            )

        conn.commit()
        log.info("All data inserted successfully")

    except psycopg2.Error as e:
        conn.rollback()
        log.error(f"Data insertion failed: {e}")
        raise
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# Data quality report
# ---------------------------------------------------------------------------

def generate_report(conn, films, festivals, directors, countries, genres, screenings, duplicates):
    """Generate a data quality report and write it to file."""
    log.info("Generating data quality report...")
    cur = conn.cursor()
    lines = []

    def section(title):
        lines.append('')
        lines.append('=' * 70)
        lines.append(title)
        lines.append('=' * 70)

    lines.append('NETWORK EFFECTS - DATA QUALITY REPORT')
    lines.append(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append(f'Source: {os.path.basename(EXCEL_PATH)}')

    # --- Record counts ---
    section('1. RECORD COUNTS')
    tables = ['films', 'festivals', 'screenings', 'directors', 'countries', 'genres',
              'film_directors', 'film_countries', 'film_genres']
    for table in tables:
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        count = cur.fetchone()[0]
        lines.append(f'  {table:20s}: {count:,}')

    # --- Date range ---
    section('2. DATE RANGE COVERAGE')
    cur.execute("SELECT MIN(year), MAX(year) FROM films WHERE year IS NOT NULL")
    min_year, max_year = cur.fetchone()
    lines.append(f'  Film production years: {min_year} - {max_year}')

    cur.execute("SELECT MIN(event_year), MAX(event_year) FROM screenings WHERE event_year IS NOT NULL")
    min_ey, max_ey = cur.fetchone()
    lines.append(f'  Screening event years: {min_ey} - {max_ey}')

    cur.execute("""
        SELECT event_year, COUNT(*) as cnt
        FROM screenings WHERE event_year IS NOT NULL
        GROUP BY event_year ORDER BY event_year
    """)
    lines.append('')
    lines.append('  Screenings per year:')
    for year, cnt in cur.fetchall():
        bar = '#' * (cnt // 10)
        lines.append(f'    {year}: {cnt:4d} {bar}')

    # --- Films with incomplete metadata ---
    section('3. FILMS WITH INCOMPLETE METADATA')
    cur.execute("""
        SELECT title, year,
            CASE WHEN duration IS NULL THEN 'missing duration' ELSE NULL END,
            CASE WHEN synopsis IS NULL OR synopsis = '' THEN 'missing synopsis' ELSE NULL END
        FROM films
        WHERE duration IS NULL OR synopsis IS NULL OR synopsis = ''
        ORDER BY title
    """)
    incomplete = cur.fetchall()
    lines.append(f'  Total films with missing data: {len(incomplete)}')
    if incomplete:
        lines.append('')
        for title, year, dur_issue, syn_issue in incomplete[:30]:
            issues = ', '.join(filter(None, [dur_issue, syn_issue]))
            lines.append(f'    "{title}" ({year}): {issues}')
        if len(incomplete) > 30:
            lines.append(f'    ... and {len(incomplete) - 30} more')

    # --- Top 10 countries by film production ---
    section('4. TOP 10 COUNTRIES BY FILM PRODUCTION')
    cur.execute("""
        SELECT c.country_name, c.country_code, COUNT(fc.film_id) as film_count
        FROM countries c
        JOIN film_countries fc ON c.country_id = fc.country_id
        GROUP BY c.country_name, c.country_code
        ORDER BY film_count DESC
        LIMIT 10
    """)
    for name, code, count in cur.fetchall():
        code_str = f' ({code})' if code else ''
        lines.append(f'  {name}{code_str:6s}: {count:4d} films')

    # --- Top 10 festivals by unique film count ---
    section('5. TOP 10 FESTIVALS BY UNIQUE FILM COUNT')
    cur.execute("""
        SELECT f.festival_name, f.event_type,
               COUNT(DISTINCT s.film_id) as film_count,
               MIN(s.event_year) as first_year,
               MAX(s.event_year) as last_year
        FROM festivals f
        JOIN screenings s ON f.festival_id = s.festival_id
        GROUP BY f.festival_name, f.event_type
        ORDER BY film_count DESC
        LIMIT 10
    """)
    for name, etype, count, fy, ly in cur.fetchall():
        lines.append(f'  {name} [{etype}] ({fy}-{ly}): {count} films')

    # --- Films with most festival screenings (top 20) ---
    section('6. TOP 20 FILMS BY FESTIVAL SCREENINGS')
    cur.execute("""
        SELECT f.title, f.year,
               COUNT(s.screening_id) as screening_count,
               COUNT(DISTINCT s.festival_id) as festival_count,
               STRING_AGG(DISTINCT d.director_name, ', ') as directors
        FROM films f
        JOIN screenings s ON f.film_id = s.film_id
        LEFT JOIN film_directors fd ON f.film_id = fd.film_id
        LEFT JOIN directors d ON fd.director_id = d.director_id
        GROUP BY f.film_id, f.title, f.year
        ORDER BY screening_count DESC
        LIMIT 20
    """)
    for i, (title, year, sc, fc, dirs) in enumerate(cur.fetchall(), 1):
        dirs_str = f' dir: {dirs}' if dirs else ''
        lines.append(f'  {i:2d}. "{title}" ({year}) - {sc} screenings at {fc} festivals{dirs_str}')

    # --- Status distribution ---
    section('7. SCREENING STATUS DISTRIBUTION')
    cur.execute("""
        SELECT status, COUNT(*) as cnt
        FROM screenings
        GROUP BY status
        ORDER BY cnt DESC
    """)
    for status, cnt in cur.fetchall():
        lines.append(f'  {status or "(null)":25s}: {cnt:4d}')

    # --- Genre distribution ---
    section('8. GENRE DISTRIBUTION')
    cur.execute("""
        SELECT g.genre_name, COUNT(fg.film_id) as film_count
        FROM genres g
        JOIN film_genres fg ON g.genre_id = fg.genre_id
        GROUP BY g.genre_name
        ORDER BY film_count DESC
    """)
    for name, count in cur.fetchall():
        lines.append(f'  {name:30s}: {count:4d} films')

    # --- Potential duplicates ---
    section('9. POTENTIAL DUPLICATE FILMS (for manual review)')
    if duplicates:
        lines.append(f'  Found {len(duplicates)} potential duplicate pairs:')
        lines.append('')
        for dup in duplicates:
            lines.append(f'  - "{dup["film_a_title"]}" vs "{dup["film_b_title"]}"')
            lines.append(f'    Year: {dup["year"]}, Title similarity: {dup["title_similarity"]}%')
            lines.append(f'    Shared directors: {", ".join(dup["shared_directors"])}')
            lines.append('')
    else:
        lines.append('  No potential duplicates detected.')

    # --- Event type distribution ---
    section('10. EVENT TYPE DISTRIBUTION')
    cur.execute("""
        SELECT f.event_type, COUNT(DISTINCT f.festival_id) as festival_count,
               COUNT(s.screening_id) as screening_count
        FROM festivals f
        LEFT JOIN screenings s ON f.festival_id = s.festival_id
        GROUP BY f.event_type
        ORDER BY screening_count DESC
    """)
    for etype, fc, sc in cur.fetchall():
        lines.append(f'  {etype or "(null)":15s}: {fc:4d} festivals, {sc:4d} screenings')

    lines.append('')
    lines.append('=' * 70)
    lines.append('END OF REPORT')
    lines.append('=' * 70)

    report_text = '\n'.join(lines)

    # Write report
    os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report_text)

    log.info(f"Report written to {REPORT_PATH}")
    cur.close()

    return report_text


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log.info("=" * 60)
    log.info("Network Effects - Film Festival Data Import")
    log.info("=" * 60)

    # 1. Read Excel data
    rows = read_excel()

    # 2. Extract normalised entities
    films, festivals, directors, countries, genres, screenings = extract_entities(rows)

    # 3. Detect potential duplicates
    duplicates = detect_duplicates(films)

    # 4. Connect to database
    conn = get_connection()

    try:
        # 5. Create schema (drops existing tables first)
        drop_all_tables(conn)
        create_schema(conn)

        # 6. Insert data
        insert_data(conn, films, festivals, directors, countries, genres, screenings)

        # 7. Generate data quality report
        report = generate_report(conn, films, festivals, directors, countries, genres, screenings, duplicates)

        # Print summary to console (handle non-ASCII characters on Windows)
        import io, sys as _sys
        _sys.stdout = io.TextIOWrapper(_sys.stdout.buffer, encoding='utf-8', errors='replace')
        print('\n' + report)

    except Exception as e:
        log.error(f"Import failed: {e}")
        raise
    finally:
        conn.close()
        log.info("Database connection closed")

    log.info("Import completed successfully!")


if __name__ == '__main__':
    main()
